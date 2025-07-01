import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os
from datetime import datetime

# --- Normalization Function ---
def normalize(text):
    if not isinstance(text, str): return ""
    text = text.upper().strip()
    text = text.replace("\xa0", " ").replace(" ", " ")
    
    text = text.replace(".", "").replace(",", "")
    text = text.replace("LIMITED", "LTD")  # Normalize "LIMITED" to "LTD"
    return text

# --- Load Excel and get original class column names ---
input_file = sys.argv[1] if len(sys.argv) > 1 else "combined_vahan_data(1).xlsx"
wb_orig = load_workbook(input_file)
ws_orig = wb_orig.active

# Get vehicle class names from D3:CA3
vehicle_class_names = [cell.value for cell in ws_orig[3][3:]]  # D3:CA3 (zero-based index 3 onwards)

# Compose full column names: [State, Month, Maker, ...vehicle class names...]
col_names = [ws_orig['A3'].value, ws_orig['B3'].value, ws_orig['C1'].value] + vehicle_class_names

# Read the data, skipping the first 3 rows (so data starts from row 4)
df = pd.read_excel(input_file, header=None, skiprows=3)
df = df.iloc[:, :len(col_names)]  # Only keep as many columns as we have names for

# Set the column names
col_names = [c if c is not None else f"COL_{i+1}" for i, c in enumerate(col_names)]
df.columns = col_names

# --- Check for 'MAKER' column ---
maker_col = next((col for col in df.columns if isinstance(col, str) and "MAKER" in col.upper()), None)
if maker_col is None:
    print("[ERROR] 'Maker' column not found in headers.")
    print("Available columns after setting custom headers:")
    for col in df.columns:
        print(f"- {col}")
    raise SystemExit(1)

# --- Clean and Normalize 'MAKER' Column ---
df = df[df[maker_col].notna() & (df[maker_col].astype(str).str.strip() != "") & (df[maker_col].astype(str).str.upper() != "MAKER")]
df[maker_col] = df[maker_col].apply(normalize)

# --- Define Filter List ---
raw_makers = [
    "Ashok Leyland Ltd", "Ather Energy Limited", "Atul Auto Ltd", "Bajaj Auto Ltd", "Baxy Limited",
    "BMW India Pvt Ltd", "Cummins India Ltd", "Daimler India Commercial Vehicles Pvt Ltd",
    "Fiat India Automobiles Pvt Ltd", "Force Motors Ltd", "Foton Motors Manufacturing India Pvt Ltd",
    "GREAVES ELECTRIC MOBILITY PVT LTD", "Hero MotoCorp Ltd", "Honda Cars India Ltd",
    "HONDA MOTORCYCLE AND SCOOTER INDIA (P) LTD", "Hyundai Motor India Limited",
    "HONDA MOTORCYCLE AND SCOOTER INDIA (P) LTD", "India Yamaha Motor Pvt Ltd", "Isuzu Motors India Pvt Ltd",
    "Jaguar Land Rover India Ltd", "JBM Auto Ltd", "JSW MG MOTOR INDIA PVT LTD",
    "Kia India Private Limited", "Mahindra & Mahindra Ltd", "Maruti Suzuki India Ltd",
    "MERCEDES-BENZ INDIA PVT LTD", "Nissan Motor India Pvt Ltd", "Okinawa Autotech Pvt. Ltd.",
    "Olectra Greentech Ltd", "PCA AUTOMOBILES INDIA PVT LTD", "Piaggio Vehicles Pvt Ltd",
    "Pinnacle Mobility Solutions Pvt Ltd", "PMI ELECTRO MOBILITY SOLUTIONS PRIVATE LIMITED",
    "Renault India Pvt Ltd", "ROYAL-ENFIELD (UNIT OF EICHER LTD)",
    "Scania Commercial Vehicles India Pvt Ltd", "Simpson & Co. Ltd",
    "Skoda Auto Volkswagen India Pvt. Ltd.", "SML ISUZU Ltd", "Suzuki Motorcycle India Pvt Ltd",
    "Switch Mobility Automotive Ltd.", "Tata Motors Ltd", "TI Clean Mobility Pvt Ltd",
    "Toyota Kirloskar Motor Pvt Ltd", "Triumph Motorcycles (India) Pvt Ltd",
    "TVS Motor Company Ltd", "VE Commercial Vehicles Ltd", "Volvo Auto India Pvt Ltd"
]
requested_makers = [normalize(m) for m in raw_makers]

# --- Match ---
available_makers = df[maker_col].unique().tolist()
matched_makers = sorted(set(requested_makers) & set(available_makers))
missing_makers = sorted(set(requested_makers) - set(available_makers))

filtered_df = df[df[maker_col].isin(matched_makers)]
non_filtered_df = df[~df[maker_col].isin(matched_makers)]

# --- Copy column widths and blank row structure from original ---
def copy_format_and_write(df_out, output_path):
    # Use the original workbook for column widths
    col_widths = {col: ws_orig.column_dimensions[col].width for col in ws_orig.column_dimensions}
    # Find blank row pattern (rows with all empty values)
    blank_rows = [i for i, row in enumerate(ws_orig.iter_rows(values_only=True), 1)
                  if all(cell is None or str(cell).strip() == '' for cell in row)]
    # Write filtered data to new workbook
    wb = Workbook()
    ws = wb.active
    # Set column widths
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
        ws.append(row)
        # Insert blank row if the original had a blank row at this position
        if r_idx in blank_rows:
            ws.append([None]*len(row))
    wb.save(output_path)

input_base = os.path.basename(input_file)
filtered_outfile = f"filtered_{input_base}"
non_filtered_outfile = f"non_filtered_{input_base}"

copy_format_and_write(filtered_df, filtered_outfile)
copy_format_and_write(non_filtered_df, non_filtered_outfile)

# --- Summary Log ---
with open("filter_log.txt", "w", encoding="utf-8") as f:
    f.write(f"✅ Matched Makers ({len(matched_makers)}):\n" + "\n".join(matched_makers) + "\n\n")
    f.write(f"❌ Missing Makers ({len(missing_makers)}):\n" + "\n".join(missing_makers) + "\n\n")

print(f"[OK] Filtering complete. Files created:\n- {filtered_outfile}\n- {non_filtered_outfile}\n- filter_log.txt")
